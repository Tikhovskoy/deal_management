import csv
import io

import openpyxl
from django.http import HttpResponse


class ContactExportService:
    def __init__(self, bitrix_token):
        self.but = bitrix_token

    def get_contacts(self):
        return list(self.but.call_list_fast(
            "crm.contact.list",
            {"select": ["ID", "NAME", "LAST_NAME", "COMPANY_ID", "PHONE", "EMAIL"]},
        ))

    def get_company_map(self):
        companies_data = self.but.call_list_fast(
            "crm.company.list", {"select": ["ID", "TITLE"]}
        )
        return {str(c["ID"]): c["TITLE"] for c in companies_data}

    def export_to_xlsx(self, contacts, company_map):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Contacts"
        header = ["имя", "фамилия", "номер телефона", "почта", "компания"]
        sheet.append(header)

        for contact in contacts:
            phone = contact.get("PHONE")[0]["VALUE"] if contact.get("PHONE") else ""
            email = contact.get("EMAIL")[0]["VALUE"] if contact.get("EMAIL") else ""
            company_name = company_map.get(str(contact.get("COMPANY_ID")), "")
            sheet.append([
                contact.get("NAME", ""),
                contact.get("LAST_NAME", ""),
                phone,
                email,
                company_name,
            ])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="contacts.xlsx"'
        return response

    def export_to_csv(self, contacts, company_map):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="contacts.csv"'
        response.write("\ufeff".encode("utf8"))

        writer = csv.writer(response)
        header = ["имя", "фамилия", "номер телефона", "почта", "компания"]
        writer.writerow(header)

        for contact in contacts:
            phone = contact.get("PHONE")[0]["VALUE"] if contact.get("PHONE") else ""
            email = contact.get("EMAIL")[0]["VALUE"] if contact.get("EMAIL") else ""
            company_name = company_map.get(str(contact.get("COMPANY_ID")), "")
            writer.writerow([
                contact.get("NAME", ""),
                contact.get("LAST_NAME", ""),
                phone,
                email,
                company_name,
            ])

        return response


class ContactImportService:
    def __init__(self, bitrix_token, user_id):
        self.but = bitrix_token
        self.user_id = user_id

    def parse_file(self, file):
        if file.name.endswith(".csv"):
            return self._parse_csv(file)
        elif file.name.endswith(".xlsx"):
            return self._parse_xlsx(file)
        else:
            raise ValueError("Неподдерживаемый формат файла. Пожалуйста, используйте CSV или XLSX.")

    def _parse_csv(self, file):
        decoded_file = file.read().decode("utf-8-sig")
        lines = decoded_file.strip().splitlines()
        header = [h.lower().strip() for h in lines[0].split(",")]
        rows = [dict(zip(header, line.split(","))) for line in lines[1:]]
        return header, rows

    def _parse_xlsx(self, file):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        header = [cell.value.lower().strip() for cell in sheet[1]]
        rows = [dict(zip(header, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
        return header, rows

    def validate_headers(self, header):
        required_headers = ["имя", "фамилия"]
        if not all(h in header for h in required_headers):
            raise ValueError(
                f"Ошибка в заголовках файла. Убедитесь, что файл содержит обязательные колонки: {required_headers}. "
                f"Найденные заголовки: {header}."
            )

    def get_existing_contacts(self):
        existing_contacts = self.but.call_list_fast(
            "crm.contact.list", {"select": ["PHONE", "EMAIL"]}
        )
        existing_phones = set()
        existing_emails = set()
        for contact in existing_contacts:
            if contact.get("PHONE"):
                existing_phones.add(contact["PHONE"][0]["VALUE"])
            if contact.get("EMAIL"):
                existing_emails.add(contact["EMAIL"][0]["VALUE"])
        return existing_phones, existing_emails

    def get_company_map(self):
        companies_data = self.but.call_list_fast(
            "crm.company.list", {"select": ["ID", "TITLE"]}
        )
        return {c["TITLE"].lower().strip(): str(c["ID"]) for c in companies_data}

    def prepare_batch_commands(self, rows, existing_phones, existing_emails, company_map):
        batch_cmds = []
        skipped_details = []
        for i, row in enumerate(rows):
            name = row.get("имя")
            last_name = row.get("фамилия")
            phone = row.get("номер телефона")
            email = row.get("почта")
            company_name = row.get("компания")

            if not name and not last_name:
                continue

            contact_name_for_report = f"{name or ''} {last_name or ''}".strip()

            if phone and phone in existing_phones:
                skipped_details.append(
                    f"Строка {i + 2}: Контакт '{contact_name_for_report}' пропущен (дубликат по номеру телефона: {phone})."
                )
                continue
            if email and email in existing_emails:
                skipped_details.append(
                    f"Строка {i + 2}: Контакт '{contact_name_for_report}' пропущен (дубликат по email: {email})."
                )
                continue

            fields = {
                "NAME": name,
                "LAST_NAME": last_name,
                "OPENED": "Y",
                "ASSIGNED_BY_ID": self.user_id,
            }
            if phone:
                fields["PHONE"] = [{"VALUE": phone, "VALUE_TYPE": "WORK"}]
            if email:
                fields["EMAIL"] = [{"VALUE": email, "VALUE_TYPE": "WORK"}]

            if company_name:
                company_id = company_map.get(str(company_name).lower().strip())
                if company_id:
                    fields["COMPANY_ID"] = company_id

            batch_cmds.append((f"cmd_{i}", "crm.contact.add", {"fields": fields}))

        return batch_cmds, skipped_details

    def process_batch_results(self, result, original_rows):
        success_details = []
        error_details = []
        if result:
            for cmd_id, res in result.items():
                index = int(cmd_id.split("_")[1])
                row_num = index + 2
                original_row_data = original_rows[index]
                contact_name = f"{original_row_data.get('имя', '')} {original_row_data.get('фамилия', '')}".strip()

                if res.get("error") is None:
                    success_details.append(
                        f"Строка {row_num}: Контакт '{contact_name}' успешно создан (ID: {res.get('result')})."
                    )
                else:
                    error_msg = res["error"].get(
                        "error_description", "Неизвестная ошибка"
                    )
                    error_details.append(
                        f"Строка {row_num}: Ошибка при создании контакта '{contact_name}' - {error_msg}"
                    )
        return success_details, error_details

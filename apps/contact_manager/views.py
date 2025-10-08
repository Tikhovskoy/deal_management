from django.shortcuts import render
from openpyxl.styles.fills import PatternFill

from apps.core.decorators import smart_auth
from .services import ContactExportService, ContactImportService

# Патч для openpyxl для обработки файлов, созданных не в MS Excel
original_init = PatternFill.__init__


def new_init(self, *args, **kwargs):
    kwargs.pop("extLst", None)
    original_init(self, *args, **kwargs)


PatternFill.__init__ = new_init


@smart_auth
def index(request):
    return render(request, "contact_manager/index.html")


@smart_auth
def export_contacts(request):
    if request.method == "POST":
        try:
            service = ContactExportService(request.bitrix_user_token)
            file_format = request.POST.get("format", "csv")

            contacts = service.get_contacts()
            company_map = service.get_company_map()

            if file_format == "xlsx":
                return service.export_to_xlsx(contacts, company_map)
            else:
                return service.export_to_csv(contacts, company_map)

        except Exception as e:
            return render(request, "contact_manager/export.html", {"error": str(e)})

    return render(request, "contact_manager/export.html")


@smart_auth
def import_contacts(request):
    context = {}
    if request.method == "POST":
        if not request.FILES.get("file"):
            context["error"] = "Файл не был загружен."
            return render(request, "contact_manager/import.html", context)

        file = request.FILES["file"]

        try:
            service = ContactImportService(request.bitrix_user_token, request.bitrix_user_token.user.bitrix_id)
            header, original_rows = service.parse_file(file)
            service.validate_headers(header)

            existing_phones, existing_emails = service.get_existing_contacts()
            company_map = service.get_company_map()

            batch_cmds, skipped_details = service.prepare_batch_commands(
                original_rows, existing_phones, existing_emails, company_map
            )

            context["skipped_count"] = len(skipped_details)
            context["skipped_details"] = skipped_details

            if batch_cmds:
                result = service.but.batch_api_call(batch_cmds)
                success_details, error_details = service.process_batch_results(result, original_rows)
                context["success_count"] = len(success_details)
                context["error_count"] = len(error_details)
                context["success_details"] = success_details
                context["error_details"] = error_details

        except Exception as e:
            context["error"] = f"Произошла ошибка при обработке файла: {str(e)}"

    return render(request, "contact_manager/import.html", context)
